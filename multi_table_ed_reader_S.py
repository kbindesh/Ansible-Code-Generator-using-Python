# Description   : This logic can be used for reading the ED document's and generate Ansible playbooks
# Author        : Team
# Created date  : 
# Last Modified : 

from openpyxl import load_workbook
import pandas as pd
import numpy as np
import re
import os

from ansiblegen.def_stat import read_def_stat_df, read_def_stat_df_new, create_playbook_for_each_env_new, create_dict_of_tasks_each_env_new, create_playbook_for_each_env, create_dict_of_tasks_each_env
from ansiblegen.def_line import read_def_stat_df_line, create_dict_of_tasks_each_env_line, \
    create_playbook_for_each_env_line

max_cols = 20
root_dir_name = 'eTax Ansible Configuration'
root_dir_path = 'D:\\Bindu\\PracticeLabs\\PythonProjects\\eTax\\eTax-Ansible-Configuration\\'
src_file = "C:\\Users\\Dell\\Documents\\ED_RHEL_English.xlsx"  # ED File path
wb = load_workbook(filename=src_file)
sheet = wb.active


class EDDocDataframe:
    def __init__(self, st_row, dim, sp, ed_df, env_list, sheet):
        self.start_row = st_row
        self.dimension = dim
        self.spectype = sp
        self.ed_df = ed_df
        self.envlist = env_list
        self.sheet = sheet

def create_dir_structure(root_dir_path, access_mode):
    env_list = ['Honban', 'Backup', 'Kensho', 'Shiken', 'Roles']
    dir_list = ['Playbook', 'Inventory', 'Plays']

    for env in env_list:
        for dir_to_create in dir_list:
            path = os.path.join(root_dir_path,env,dir_to_create)
            print('====='+str(path))
            if not os.path.exists(path):
                try:
                    os.makedirs(path, access_mode)
                except OSError:
                    print("Creation of the directory %s failed" % path)
            else:
                print("Directory already existed : ", path)

def get_table_envlist(row_val):
    envlist = []
    for col_index in range(2, max_cols):
        if sheet.cell(row=row_val, column=col_index).value != None and (
                'env' in sheet.cell(row=row_val, column=col_index).value):
            envlist.append(sheet.cell(row=row_val, column=col_index).value)
    return envlist


def get_table_dimentions(row_val, col_val):
    col_count = 0
    row_count = 0
    for col_index in range(2, sheet.max_column + 1):
        if sheet.cell(row=row_val, column=col_index).value != None:
            col_count = col_count + 1

    for row_index in range(row_val + 1, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=2).value != None:
            row_count = row_count + 1
        else:
            break
    print("Dataframe starts at row: " + str(row_val))
    return row_val, row_count, col_count


def create_dataframe(tabledata):
    dataframe = pd.DataFrame.from_dict(tabledata)
    return dataframe


def get_table_data(row, row_count, col_count):
    table_dict = {}
    for col in range(2, max_cols):
        col_val_list = []
        col_header = sheet.cell(row=row, column=col).value
        if sheet.cell(row=row, column=col).value != None:
            for r in range(row + 1, row + row_count + 1):
                col_val_list.append(sheet.cell(row=r, column=col).value)
            table_dict[col_header] = col_val_list

    return table_dict


def main():
    table_dimensions = ()
    master_df_list = []
    stat_df_list = []
    dir_df_list = []
    cmd_df_list = []
    line_df_list = []
    stat_master_df = pd.DataFrame()
    dir_master_df = pd.DataFrame()
    cmd_master_df = pd.DataFrame()
    line_master_df = pd.DataFrame()
    is_multiple_stat = is_multiple_dir = is_multiple_cmd = is_multiple_line = False

    create_dir_structure(root_dir_path, 0o755)

    for row_index in range(1, sheet.max_row + 1):
        spec_type = sheet.cell(row=row_index, column=1).value
        if spec_type != None:
            spec = spec_type[4:]
            print("SpecType : " + str(spec))
            table_dimensions = get_table_dimentions(row_index, 1)
            print("Table Dimension : " + str(table_dimensions))

            table_data = get_table_data(*table_dimensions)
            dataframe = create_dataframe(table_data)
            dataframe.fillna(value=np.nan, inplace=True)
            env_list = get_table_envlist(row_index)
            if spec == 'stat':
                if is_multiple_stat == False:
                    stat_master_df = stat_master_df.append(dataframe)
                    stat_df_list.append(
                        EDDocDataframe(table_dimensions[0], table_dimensions, spec, stat_master_df, env_list, sheet))
                    is_multiple_stat = True
                else:
                    stat_master_df = stat_master_df.append(dataframe)
                    stat_df_list.clear()
                    stat_df_list.append(
                        EDDocDataframe(table_dimensions[0], table_dimensions, spec, stat_master_df, env_list, sheet))
            elif spec == 'dir':
                if is_multiple_dir:
                    dir_master_df = dir_master_df.append(dataframe)
                    dir_df_list.append(
                        EDDocDataframe(table_dimensions[0], table_dimensions, spec, dir_master_df, env_list, sheet))
                    is_multiple_dir = True
                else:
                    dir_master_df = dir_master_df.append(dataframe)
                    dir_df_list.clear()
                    dir_df_list.append(
                        EDDocDataframe(table_dimensions[0], table_dimensions, spec, dir_master_df, env_list, sheet))
            elif spec == 'cmd':
                if is_multiple_cmd:
                    cmd_master_df = dir_master_df.append(dataframe)
                    cmd_df_list.append(
                        EDDocDataframe(table_dimensions[0], table_dimensions, spec, cmd_master_df, env_list, sheet))
                    is_multiple_cmd = True
                else:
                    cmd_master_df = dir_master_df.append(dataframe)
                    cmd_df_list.clear()
                    cmd_df_list.append(
                        EDDocDataframe(table_dimensions[0], table_dimensions, spec, cmd_master_df, env_list, sheet))
            elif spec == 'line':
                if is_multiple_line:
                    line_master_df = line_master_df.append(dataframe)
                    line_df_list.append(
                        EDDocDataframe(table_dimensions[0], table_dimensions, spec, line_master_df, env_list, sheet))
                    is_multiple_cmd = True
                else:
                    line_master_df = line_master_df.append(dataframe)
                    cmd_df_list.clear()
                    line_df_list.append(
                        EDDocDataframe(table_dimensions[0], table_dimensions, spec, line_master_df, env_list, sheet))
            print("=====================Table_dataframe=======================\n")

    # Master DF list with one consolidated DF of every def-type | NaN will be filled for the values not present

    master_df_list = stat_df_list + dir_df_list + cmd_df_list + line_df_list

    # PROCESSING SECTION : ref class EDDocDataframe for other attributes

    for df in master_df_list:
        if df.spectype == 'stat':
            stat_vars_list, symlnk_vars_list = read_def_stat_df_new(df.ed_df)
            play_vars = create_dict_of_tasks_each_env_new(stat_vars_list, symlnk_vars_list)
            create_playbook_for_each_env_new(play_vars)
        if df.spectype == 'dir':
            # dictFilePermission = read_def_dir_df(df.ed_df)
            # taskDict = create_dict_of_tasks_each_env(dictFilePermission)
            # create_playbook_for_each_env(taskDict)
            pass
        if df.spectype == 'cmd':
            # dictFilePermission = read_def_cmd_df(df.ed_df)
            # taskDict = create_dict_of_tasks_each_env(dictFilePermission)
            # create_playbook_for_each_env(taskDict)
            pass


if __name__ == '__main__':
    main()

