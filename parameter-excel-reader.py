from openpyxl import load_workbook
import pandas as pd

src_file = "C:\\Users\\DELL\\Documents\\multi-dataframe-excel.xlsx"

def get_table_dimentions(row_val, col_val):
    col_count = 0
    row_count = 0
    for col_index in range(1,sheet.max_column+1):
        if sheet.cell(row=row_val, column=col_index).value != None:
            col_count = col_count +1

    for row_index in range(row_val, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=2).value != None:
            row_count = row_count +1
        else:
            break
    print("Dataframe starts at row: " +str(row_val))
    print("Max Col : " +str(col_count))
    print ("Max Rows : " +str(row_count))
    return row_val,row_count,col_count

def create_dataframe(tabledata):
    dataframe = pd.DataFrame.from_dict(tabledata)
    return dataframe

def get_table_data(row,row_count,col_count):
    table_dict = {}
    for col in range(2, col_count+1):
        col_val_list = []
        for r in range(row+1,row+row_count):
            col_header = sheet.cell(row=row, column=col).value
            col_val_list.append(sheet.cell(row=r, column=col).value)
        table_dict[col_header] = col_val_list

    return table_dict


wb = load_workbook(filename = src_file)
print(wb.sheetnames)

sheet = wb.active
table_dimensions = ()
for row_index in range(1, sheet.max_row+1):
    spec_type = sheet.cell(row=row_index, column=1).value
    if spec_type != None:
        spec = spec_type[4:]

        table_dimensions = get_table_dimentions(row_index, 1)  # 1 is column index; it starts from 1st col
        print("Table Dimension : " + str(table_dimensions))

        table_data = get_table_data(*table_dimensions)
        print("Dictionary : " + str(table_data))

        table_dataframe = create_dataframe(table_data)
        print("Data Frame :\n" + str(table_dataframe))

        if spec == 'line' or spec == 'stat':
            pass
