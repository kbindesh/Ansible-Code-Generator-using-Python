# Description   : This logic can be used for reading the ED document's active sheet and get the table dimentions and its data  
# Author        : 
# Created date  : 16-Apr-2021
# Last Modified : 16-Apr-2021

import os
from openpyxl import load_workbook
import pandas as pd
import re
from ansiblegen.def_stat import read_def_stat_df, create_playbook_for_each_env, create_dict_of_tasks_each_env
#from ansiblegen.def_line import  read_def_line_df

max_cols = 20
src_file = "C:\\Users\\DELL\\Documents\\ED_RHEL_English.xlsx"
wb = load_workbook(filename=src_file)
sheet = wb.active

class EDDocDataframe:
    def __init__(self,st_row,dim,sp,ed_df,env_list):
        self.start_row = st_row
        self.dimension = dim
        self.spectype = sp
        self.ed_df = ed_df
        self.envlist = env_list

def get_table_envlist(row_val):
    envlist = []
    for col_index in range(2, max_cols + 3):
        if sheet.cell(row=row_val, column=col_index).value != None and (
                'env' in sheet.cell(row=row_val, column=col_index).value):
            envlist.append(sheet.cell(row=row_val, column=col_index).value)
    return envlist

def get_table_dimentions(row_val, col_val):
    col_count = 0
    row_count = 0
    for col_index in range(2, sheet.max_column + 1):
        if sheet.cell(row=row_val, column=col_index).value != None:
            col_count = col_count + 1

    for row_index in range(row_val + 1, sheet.max_row + 1):
        if sheet.cell(row=row_index, column=2).value != None:
            row_count = row_count + 1
        else:
            break
    print("Dataframe starts at row: " + str(row_val))
    return row_val, row_count, col_count


def create_dataframe(tabledata):
    dataframe = pd.DataFrame.from_dict(tabledata)
    return dataframe

def get_table_data(row, row_count, col_count):
    table_dict = {}
    for col in range(2, col_count + 3):
        col_val_list = []
        col_header = sheet.cell(row=row, column=col).value
        if sheet.cell(row=row, column=col).value != None:
            for r in range(row + 1, row + row_count + 1):
                col_val_list.append(sheet.cell(row=r, column=col).value)
            table_dict[col_header] = col_val_list

    return table_dict

def df_to_dict(dataframe):
    df_dict = {}
    return dataframe.to_dict(orient='dict')

def main():
    table_dimensions = ()
    isMultiBlock = False
    final_stat_dict = {}
    df_list = []
    for row_index in range(1, sheet.max_row + 1):
        spec_type = sheet.cell(row=row_index, column=1).value
        if spec_type != None:
            spec = spec_type[4:]
            print("SpecType : "+str(spec))
            table_dimensions = get_table_dimentions(row_index, 1)
            print("Table Dimension : " + str(table_dimensions))

            env_list = get_table_envlist(row_index)
            table_data = get_table_data(*table_dimensions)
            stat_dataframe = create_dataframe(table_data)
            df_list.append(EDDocDataframe(table_dimensions[0],table_dimensions,spec,stat_dataframe,env_list))
            print("=====================table_dataframe=======================\n")

    for df in df_list:
        if df.spectype == 'stat':
            if isMultiBlock == False:
                stat_df = df.ed_df
                isMultiBlock = True
            else:
                stat_df = pd.concat([stat_df,df.ed_df])             # Merged DF
                dictFilePermission = read_def_stat_df(stat_df)
                taskDict = create_dict_of_tasks_each_env(dictFilePermission)
                create_playbook_for_each_env(taskDict)
        if df.spectype == 'dir':
            pass
        if df.spectype == 'line':
            # Will place the Akshay's logic here for line dfs
            pass
        # List of
        # CREATION OF FILE PERMISSION DICTIONARIES

if __name__ == '__main__':
    main()
